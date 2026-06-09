import os
import sys
import re
import json
import time
import html
import random
import sqlite3
import subprocess
from datetime import datetime, UTC
from pathlib import Path
from urllib.parse import urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from config.settings import *
from scraper.listings import collect_listings
from processing.utils import clean_text
from scraper.website import *
from scraper.profiles import enrich_rows

# --------------------------------------------------
# CONFIG
# --------------------------------------------------

PRODUCT_HIERARCHY = {
    "Beef": {
        "Beef": ["beef"],
        "Grainfed Beef": ["grainfed beef", "grain-fed beef"],
        "Grassfed Beef": ["grassfed beef", "grass-fed beef"],
        "Wagyu Beef": ["wagyu beef", "wagyu"],
        "Beef Offal": ["beef offal"],
        "Value Added Beef": [
            "value added beef",
            "value-added beef",
            "value added beef products",
            "value-added beef products",
        ],
        "Veal": ["veal"],
    },
    "Lamb and Sheepmeat": {
        "Lamb": ["lamb"],
        "Mutton": ["mutton"],
        "Sheepmeat": ["sheepmeat"],
        "Sheep Offal": ["sheep offal"],
        "Value Added Lamb/Mutton": [
            "value added lamb",
            "value added lamb/mutton",
            "value-added lamb",
            "value-added lamb/mutton",
            "value added lamb/mutton products",
            "value-added lamb/mutton products",
        ],
    },
    "Goat": {
        "Goatmeat": ["goatmeat", "goat meat"],
        "Goat Offal": ["goat offal"],
        "Value Added Goatmeat": [
            "value added goatmeat",
            "value-added goatmeat",
            "value added goatmeat products",
            "value-added goatmeat products",
        ],
    },
}

ABN_NEARBY_PATTERN = re.compile(r"\bABN\b", re.IGNORECASE)
YEAR_RANGE_PATTERN = re.compile(r"^\d{4}\s*[-–]\s*\d{4}$")
BARE_DOMAIN_PATTERN = re.compile(
    r"\b(?:[A-Za-z0-9-]+\.)+[A-Za-z]{2,}(?:/[^\s]*)?\b"
)

# --------------------------------------------------
# GENERAL HELPERS
# --------------------------------------------------

def extract_between(text, start_label, end_labels):
    text = clean_text(text)
    start_match = re.search(
        re.escape(start_label) + r"\s+(.*)", text, flags=re.IGNORECASE
    )
    if not start_match:
        return ""

    remainder = start_match.group(1)
    end_positions = []

    for label in end_labels:
        m = re.search(
            r"\s" + re.escape(label) + r"\b",
            remainder,
            flags=re.IGNORECASE,
        )
        if m:
            end_positions.append(m.start())

    if end_positions:
        return clean_text(remainder[:min(end_positions)])

    return clean_text(remainder)


def flatten_hierarchy_to_product_map():
    out = {}
    for family, products in PRODUCT_HIERARCHY.items():
        for product_name, keywords in products.items():
            out[product_name] = {"family": family, "keywords": keywords}
    return out


def normalize_phone(phone):
    phone = clean_text(phone)
    phone = re.sub(r"[^\d+()\-\ ]", "", phone)
    phone = re.sub(r"\s+", " ", phone).strip()
    return phone


def is_probable_abn(candidate, surrounding_text=""):
    digits = re.sub(r"\D", "", candidate)
    return len(digits) == 11 and bool(ABN_NEARBY_PATTERN.search(surrounding_text or ""))


def filter_phone_candidates(candidates, surrounding_text=""):
    cleaned = []
    for raw in candidates:
        candidate = normalize_phone(raw)
        digits = re.sub(r"\D", "", candidate)
        if not candidate:
            continue
        if YEAR_RANGE_PATTERN.match(candidate):
            continue
        if len(digits) < 8:
            continue
        if is_probable_abn(candidate, surrounding_text):
            continue
        cleaned.append(candidate)
    return unique_nonblank(cleaned)


def build_field_texts_for_matching(company_row):
    return {
        "description": clean_text(company_row.get("description", "")),
        "page_heading": clean_text(company_row.get("page_heading", "")),
        "meta_description": clean_text(company_row.get("meta_description", "")),
        "details_json": clean_text(company_row.get("details_json", "")),
        "page_text_excerpt": clean_text(company_row.get("page_text_excerpt", "")),
    }


def score_product_matches(field_texts, product_map):
    matches = []

    for source_field, source_text in field_texts.items():
        lower_text = source_text.lower()
        if not lower_text:
            continue

        for product_name, meta in product_map.items():
            for keyword in meta["keywords"]:
                if keyword.lower() in lower_text:
                    confidence = FIELD_WEIGHTS.get(source_field, 0.50)
                    matches.append({
                        "product_name": product_name,
                        "family_name": meta["family"],
                        "matched_keyword": keyword,
                        "source_field": source_field,
                        "source_excerpt": source_text[:250],
                        "match_confidence": confidence,
                    })
                    break

    return matches


def extract_countries_from_details(details):
    chunks = []
    for k, v in details.items():
        key = clean_text(k).lower()
        if any(token in key for token in COUNTRY_FIELD_KEYWORDS):
            chunks.append(clean_text(v))
    return " | ".join(unique_nonblank(chunks))


def has_any_keyword(text, keywords):
    text = clean_text(text).lower()
    return any(keyword.lower() in text for keyword in keywords)


def compute_brokerage_signals(row):
    website = clean_text(row.get("website"))
    emails = clean_text(row.get("emails"))
    phones = clean_text(row.get("phones"))
    countries_served = clean_text(row.get("countries_served"))
    certifications = clean_text(row.get("certifications"))
    accreditations = clean_text(row.get("accreditations"))
    product_families = clean_text(row.get("product_families"))
    product_variants = clean_text(row.get("product_variants"))

    score = 0
    if emails:
        score += 5
    if website:
        score += 3
    if phones:
        score += 2
    if certifications:
        score += 3
    if accreditations:
        score += 3

    product_text = f"{product_families} {product_variants}"
    if has_any_keyword(product_text, PREMIUM_PRODUCT_KEYWORDS):
        score += 3
    if has_any_keyword(product_text, SINGAPORE_FOCUS_PRODUCTS):
        score += 2
    if TARGET_COUNTRY.lower() in countries_served.lower():
        score += 4

    if score >= 14:
        priority = "High"
    elif score >= 8:
        priority = "Medium"
    else:
        priority = "Low"

    outreach_ready = "Yes" if (website and (emails or phones)) else "No"
    singapore_fit = "Yes" if TARGET_COUNTRY.lower() in countries_served.lower() else "No"
    premium_beef_fit = "Yes" if has_any_keyword(product_text, PREMIUM_PRODUCT_KEYWORDS) else "No"

    checks = []
    if not countries_served:
        checks.append("Confirm countries serviced")
    if TARGET_COUNTRY.lower() not in countries_served.lower():
        checks.append("Confirm Singapore service capability")
    if not accreditations:
        checks.append("Confirm market accreditations")
    if not certifications:
        checks.append("Confirm quality/certification status")
    if not emails and not phones:
        checks.append("Manual contact research needed")

    return {
        "brokerage_fit_score": score,
        "priority_band": priority,
        "outreach_ready": outreach_ready,
        "singapore_fit": singapore_fit,
        "premium_beef_fit": premium_beef_fit,
        "manual_checks_needed": "; ".join(checks),
    }

# --------------------------------------------------
# EXCEL WRITING / DASHBOARD STYLING
# --------------------------------------------------

def write_styled_excel(path, sheets):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = {
            str(cell.value).strip().lower(): idx + 1
            for idx, cell in enumerate(ws[1])
            if cell.value is not None
        }

        for url_col_name in ["profile url", "image url", "website", "profile_url", "image_url"]:
            col_idx = headers.get(url_col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        cell.hyperlink = str(cell.value)
                        cell.style = "Hyperlink"

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)

    wb.save(path)


def dashboard_polish_workbook(path):
    wb = load_workbook(path)

    desired_order = [
        "Summary",
        "Brokerage Targets",
        "Singapore Focus",
        "Outreach Ready",
        "Manual Verification",
        "Product Summary",
        "Certification Summary",
        "Accreditation Summary",
        "State Summary",
        "Exporter Type Summary",
        "Missing Data",
        "Company Profile",
        "Company Products",
        "Companies",
        "Contacts",
        "Product Match Audit",
        "Certifications",
        "Accreditations",
        "Product Hierarchy",
    ]
    ordered = [wb[s] for s in desired_order if s in wb.sheetnames] + [wb[s] for s in wb.sheetnames if s not in desired_order]
    wb._sheets = ordered

    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    blue_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    pink_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws.sheet_view.showGridLines = False
        for row in range(2, ws.max_row + 1):
            metric = ws.cell(row=row, column=1).value
            val_cell = ws.cell(row=row, column=2)
            if metric and metric != "Generated UTC":
                val_cell.fill = green_fill
                val_cell.font = Font(bold=True)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 22

    for sheet_name in [
        "Product Summary",
        "Certification Summary",
        "Accreditation Summary",
        "State Summary",
        "Exporter Type Summary",
        "Brokerage Targets",
        "Singapore Focus",
        "Outreach Ready",
    ]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_view.showGridLines = False
            if ws.max_row >= 2 and ws.max_column >= 2:
                headers = {
                    str(ws.cell(row=1, column=i).value or "").strip().lower(): i
                    for i in range(1, ws.max_column + 1)
                }
                score_col = headers.get("brokerage_fit_score") or 2
                if score_col <= ws.max_column:
                    rng = f"{get_column_letter(score_col)}2:{get_column_letter(score_col)}{ws.max_row}"
                    ws.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="min",
                            start_color="F8696B",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FFEB84",
                            end_type="max",
                            end_color="63BE7B",
                        ),
                    )

    for sheet_name in [
        "Company Profile",
        "Company Products",
        "Companies",
        "Brokerage Targets",
        "Singapore Focus",
        "Outreach Ready",
        "Manual Verification",
    ]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = {
                str(ws.cell(row=1, column=i).value or "").strip().lower(): i
                for i in range(1, ws.max_column + 1)
            }

            for key in [
                "name",
                "company_name",
                "exporter_type",
                "product_families",
                "product_variants",
                "certifications",
                "accreditations",
                "state",
                "countries_served",
                "brokerage_fit_score",
                "priority_band",
                "singapore_fit",
                "outreach_ready",
            ]:
                idx = headers.get(key)
                if idx:
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=idx).fill = blue_fill

            for text_col in ["page_text_preview", "meta_description", "description", "manual_checks_needed"]:
                idx = headers.get(text_col)
                if idx:
                    ws.column_dimensions[get_column_letter(idx)].width = 40

    if "Contacts" in wb.sheetnames:
        ws = wb["Contacts"]
        ws.sheet_view.showGridLines = False
        for row in range(2, ws.max_row + 1):
            email = str(ws.cell(row=row, column=2).value or "").strip()
            phone = str(ws.cell(row=row, column=3).value or "").strip()
            if email or phone:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill

    if "Missing Data" in wb.sheetnames:
        ws = wb["Missing Data"]
        ws.sheet_view.showGridLines = False
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = pink_fill

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.topLeftCell = "A1"

    wb.save(path)

# --------------------------------------------------
# EXCEL LOCK / FALLBACK HANDLING
# --------------------------------------------------

def cleanup_fallback_excels_if_master_written(used_fallback):
    if used_fallback:
        print("⚠️ Keeping fallback Excel files (main file was locked).")
        return

    files = list(Path(".").glob("exporters_master_locked_export_*.xlsx"))
    if not files:
        return

    for f in files:
        try:
            f.unlink()
            print(f"Deleted fallback file: {f.name}")
        except Exception as e:
            print(f"Could not delete {f.name}: {e}")

    print("All fallback Excel files cleaned up.")


def finalize_locked_excel(master_path, fallback_path, poll_seconds=3, max_attempts=1200):
    master_path = os.path.abspath(master_path)
    fallback_path = os.path.abspath(fallback_path)

    for _ in range(max_attempts):
        if not os.path.exists(fallback_path):
            return
        try:
            os.replace(fallback_path, master_path)
            for f in Path(".").glob("exporters_master_locked_export_*.xlsx"):
                try:
                    if os.path.abspath(str(f)) not in {master_path, fallback_path}:
                        f.unlink()
                except Exception:
                    pass
            return
        except PermissionError:
            time.sleep(poll_seconds)
        except Exception:
            time.sleep(poll_seconds)


def launch_finalize_in_background(master_path, fallback_path):
    python_exe = sys.executable
    script_path = os.path.abspath(__file__)

    creation_flags = 0
    if os.name == "nt":
        creation_flags = (
            getattr(subprocess, "DETACHED_PROCESS", 0)
            | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
        )

    subprocess.Popen(
        [python_exe, script_path, "--finalize-fallback", master_path, fallback_path],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        stdin=subprocess.DEVNULL,
        close_fds=True,
        creationflags=creation_flags,
    )


def safe_write_excel(path, sheets):
    try:
        write_styled_excel(path, sheets)
        return path, False
    except PermissionError:
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        fallback_path = path.replace(".xlsx", f"_locked_export_{timestamp}.xlsx")
        write_styled_excel(fallback_path, sheets)
        launch_finalize_in_background(path, fallback_path)
        return fallback_path, True

# --------------------------------------------------
# SCRAPING PROFILES / WEBSITES
# --------------------------------------------------

def clean_dataframe(df):
    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]):
            df[col] = df[col].fillna("").astype(str).map(clean_text)

    df.insert(0, "exported_at_utc", datetime.now(UTC).isoformat(timespec="seconds"))

    dedupe_cols = [c for c in ["company_name", "profile_url"] if c in df.columns]
    if dedupe_cols:
        df = df.drop_duplicates(subset=dedupe_cols, keep="first").reset_index(drop=True)

    if "company_name" in df.columns:
        sort_cols = [c for c in ["company_name", "profile_url"] if c in df.columns]
        df = df.sort_values(by=sort_cols, kind="stable").reset_index(drop=True)

    return df

# --------------------------------------------------
# DATABASE SETUP
# --------------------------------------------------

def init_schema(cur):
    cur.execute("PRAGMA foreign_keys = OFF;")

    sql = (
        "DROP VIEW IF EXISTS v_product_hierarchy;"
        "DROP VIEW IF EXISTS v_company_products;"
        "DROP VIEW IF EXISTS v_company_profile;"
        "DROP TABLE IF EXISTS product_match_audit;"
        "DROP TABLE IF EXISTS company_certifications;"
        "DROP TABLE IF EXISTS certifications;"
        "DROP TABLE IF EXISTS company_accreditations;"
        "DROP TABLE IF EXISTS accreditations;"
        "DROP TABLE IF EXISTS product_families;"
        "DROP TABLE IF EXISTS products;"
        "DROP TABLE IF EXISTS company_products;"
        "DROP TABLE IF EXISTS product_family_matches;"
        "DROP TABLE IF EXISTS contacts;"
        "DROP TABLE IF EXISTS attributes;"
        "DROP TABLE IF EXISTS companies;"

        "CREATE TABLE companies ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "name TEXT, description TEXT, profile_url TEXT UNIQUE, image_url TEXT,"
        "exporter_type TEXT, licence_number TEXT, establishment_numbers TEXT,"
        "website TEXT, address TEXT, abn TEXT, state TEXT, postcode TEXT,"
        "countries_served TEXT,"
        "page_title TEXT, page_heading TEXT, meta_description TEXT, meta_title TEXT,"
        "page_text_excerpt TEXT, details_json TEXT, profile_error TEXT);"

        "CREATE TABLE contacts ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, email TEXT, phone TEXT, contact_source TEXT,"
        "UNIQUE(company_id, email, phone, contact_source),"
        "FOREIGN KEY(company_id) REFERENCES companies(id));"

        "CREATE TABLE attributes ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, name TEXT, value TEXT,"
        "FOREIGN KEY(company_id) REFERENCES companies(id));"

        "CREATE TABLE certifications (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE);"
        "CREATE TABLE company_certifications (company_id INTEGER, certification_id INTEGER, UNIQUE(company_id, certification_id), FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(certification_id) REFERENCES certifications(id));"

        "CREATE TABLE accreditations (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE);"
        "CREATE TABLE company_accreditations (company_id INTEGER, accreditation_id INTEGER, UNIQUE(company_id, accreditation_id), FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(accreditation_id) REFERENCES accreditations(id));"

        "CREATE TABLE product_families (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, sort_order INTEGER);"
        "CREATE TABLE products (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, family_id INTEGER, parent_product_id INTEGER, hierarchy_level INTEGER, FOREIGN KEY(family_id) REFERENCES product_families(id), FOREIGN KEY(parent_product_id) REFERENCES products(id));"

        "CREATE TABLE company_products ("
        "company_id INTEGER, product_id INTEGER, match_source TEXT, source_field TEXT, matched_keyword TEXT, match_confidence REAL,"
        "UNIQUE(company_id, product_id, matched_keyword, source_field),"
        "FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(product_id) REFERENCES products(id));"

        "CREATE TABLE product_family_matches (company_id INTEGER, family_id INTEGER, UNIQUE(company_id, family_id), FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(family_id) REFERENCES product_families(id));"

        "CREATE TABLE product_match_audit ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, product_id INTEGER, family_id INTEGER, source_field TEXT, matched_keyword TEXT, source_excerpt TEXT, match_confidence REAL,"
        "FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(product_id) REFERENCES products(id), FOREIGN KEY(family_id) REFERENCES product_families(id));"
    )

    cur.executescript(sql)
    cur.execute("PRAGMA foreign_keys = ON;")


def upsert_lookup_and_link(cur, company_id, table, link_table, link_fk, value):
    cur.execute(f"INSERT OR IGNORE INTO {table} (name) VALUES (?)", (value,))
    cur.execute(f"SELECT id FROM {table} WHERE name = ?", (value,))
    item_id = cur.fetchone()[0]
    cur.execute(f"INSERT OR IGNORE INTO {link_table} (company_id, {link_fk}) VALUES (?, ?)", (company_id, item_id))


def parse_company_fields(text):
    fields = {
        "website": "",
        "address": "",
        "abn": "",
        "exporter_type": "",
        "licence_number": "",
        "establishment_numbers": "",
        "state": "",
        "postcode": "",
    }

    m = re.search(r"Website\s+(https?://\S+|www\.\S+)", text, flags=re.IGNORECASE)
    if m:
        fields["website"] = clean_text(m.group(1))

    fields["address"] = extract_between(
        text,
        "Address",
        ["ABN", "Exporter Type", "Exporter type", "Licence number", "Establishment numbers", "Accreditations", "Certifications", "Login"],
    )

    m = re.search(
        r"ABN\s+([A-Z0-9 ]+?)(?=\s+(Exporter Type|Exporter type|Licence number|Establishment numbers|Accreditations|Certifications|Login)\b)",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        fields["abn"] = clean_text(m.group(1))

    m = re.search(
        r"Exporter\s+[Tt]ype\s+(.*?)(?=\s+(Licence number|Establishment numbers|Accreditations|Certifications|Login)\b)",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        fields["exporter_type"] = clean_text(m.group(1))

    m = re.search(
        r"Licence number\s+(.*?)(?=\s+(Establishment numbers|Accreditations|Certifications|Login)\b)",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        fields["licence_number"] = clean_text(m.group(1))

    m = re.search(
        r"Establishment numbers\s+(.*?)(?=\s+(Accreditations|Certifications|Login)\b)",
        text,
        flags=re.IGNORECASE,
    )
    if m:
        fields["establishment_numbers"] = clean_text(m.group(1))

    m = re.search(r"\b(NSW|VIC|QLD|WA|SA|TAS|ACT|NT)\s+(\d{4})\b", fields["address"])
    if m:
        fields["state"] = m.group(1)
        fields["postcode"] = m.group(2)

    return fields


def build_database(df):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    init_schema(cur)

    for _, row in df.iterrows():
        cur.execute(
            """
            INSERT OR IGNORE INTO companies (
                name, description, profile_url, image_url,
                website, countries_served,
                page_title, page_heading, meta_description, meta_title,
                page_text_excerpt, details_json, profile_error
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                clean_text(row.get("company_name")),
                clean_text(row.get("description")),
                clean_text(row.get("profile_url")),
                clean_text(row.get("image_url")),
                clean_text(row.get("website")),
                clean_text(row.get("countries_served")),
                clean_text(row.get("page_title")),
                clean_text(row.get("page_heading")),
                clean_text(row.get("meta_description")),
                clean_text(row.get("meta_title")),
                clean_text(row.get("page_text_excerpt")),
                clean_text(row.get("details_json")),
                clean_text(row.get("profile_error")),
            ),
        )

        cur.execute(
            "SELECT id FROM companies WHERE profile_url = ?",
            (clean_text(row.get("profile_url")),)
        )
        company_id = cur.fetchone()[0]

        email_values = (
            unique_nonblank(clean_text(row.get("emails")).split(";"))
            if clean_text(row.get("emails")) else []
        )
        phone_values = (
            unique_nonblank(clean_text(row.get("phones")).split(";"))
            if clean_text(row.get("phones")) else []
        )

        for email in email_values:
            if email and "mla.com.au" not in email.lower():
                cur.execute(
                    "INSERT OR IGNORE INTO contacts (company_id, email, phone, contact_source) VALUES (?, ?, ?, ?)",
                    (company_id, email, "", "regex_or_href_or_website"),
                )

        for phone in phone_values:
            if phone:
                cur.execute(
                    "INSERT OR IGNORE INTO contacts (company_id, email, phone, contact_source) VALUES (?, ?, ?, ?)",
                    (company_id, "", phone, "regex_or_href_or_website"),
                )

        for raw_name in ["page_title", "page_heading", "meta_description", "meta_title", "page_text_excerpt", "details_json", "countries_served"]:
            value = clean_text(row.get(raw_name))
            if value:
                cur.execute(
                    "INSERT INTO attributes (company_id, name, value) VALUES (?, ?, ?)",
                    (company_id, raw_name, value),
                )

        text = " ".join([
            clean_text(row.get("description")),
            clean_text(row.get("page_heading")),
            clean_text(row.get("page_text_excerpt")),
            clean_text(row.get("details_json")),
        ])

        fields = parse_company_fields(text)
        final_website = clean_text(row.get("website")) or fields["website"]
        final_countries = clean_text(row.get("countries_served"))

        cur.execute(
            """
            UPDATE companies
            SET website = ?, address = ?, abn = ?, exporter_type = ?, licence_number = ?, establishment_numbers = ?, state = ?, postcode = ?, countries_served = ?
            WHERE id = ?
            """,
            (
                final_website,
                fields["address"],
                fields["abn"],
                fields["exporter_type"],
                fields["licence_number"],
                fields["establishment_numbers"],
                fields["state"],
                fields["postcode"],
                final_countries,
                company_id,
            ),
        )

        cert_text = extract_between(text, "Certifications", ["Login", "Australia", "Privacy", "View Report"])
        for cert in [c for c in KNOWN_CERTIFICATIONS if c in cert_text]:
            upsert_lookup_and_link(cur, company_id, "certifications", "company_certifications", "certification_id", cert)

        acc_text = extract_between(text, "Accreditations", ["Certifications", "Login", "Australia", "Privacy", "View Report"])
        for acc in [a for a in KNOWN_ACCREDITATIONS if re.search(r"\b" + re.escape(a) + r"\b", acc_text)]:
            upsert_lookup_and_link(cur, company_id, "accreditations", "company_accreditations", "accreditation_id", acc)

    family_order = list(PRODUCT_HIERARCHY.keys())
    for idx, family_name in enumerate(family_order, start=1):
        cur.execute(
            "INSERT OR IGNORE INTO product_families (name, sort_order) VALUES (?, ?)",
            (family_name, idx),
        )

    product_name_to_ids = {}
    family_name_to_id = {}

    for family_name, variants in PRODUCT_HIERARCHY.items():
        cur.execute("SELECT id FROM product_families WHERE name = ?", (family_name,))
        family_id = cur.fetchone()[0]
        family_name_to_id[family_name] = family_id

        cur.execute(
            "INSERT OR IGNORE INTO products (name, family_id, parent_product_id, hierarchy_level) VALUES (?, ?, NULL, 1)",
            (family_name, family_id),
        )
        cur.execute("SELECT id, family_id FROM products WHERE name = ?", (family_name,))
        root_product_id, root_family_id = cur.fetchone()
        product_name_to_ids[family_name] = (root_product_id, root_family_id)

        for product_name in variants.keys():
            if product_name == family_name:
                continue
            cur.execute(
                "INSERT OR IGNORE INTO products (name, family_id, parent_product_id, hierarchy_level) VALUES (?, ?, ?, 2)",
                (product_name, family_id, root_product_id),
            )
            cur.execute("SELECT id, family_id FROM products WHERE name = ?", (product_name,))
            product_name_to_ids[product_name] = cur.fetchone()

    product_map = flatten_hierarchy_to_product_map()

    cur.execute("""
        SELECT
            id,
            description,
            page_heading,
            meta_description,
            details_json,
            page_text_excerpt
        FROM companies
    """)
    company_rows = cur.fetchall()

    for company_id, description, page_heading, meta_description, details_json, page_text_excerpt in company_rows:
        company_row = {
            "description": description,
            "page_heading": page_heading,
            "meta_description": meta_description,
            "details_json": details_json,
            "page_text_excerpt": page_text_excerpt,
        }

        field_texts = build_field_texts_for_matching(company_row)
        matches = score_product_matches(field_texts, product_map)
        matched_families = set()

        for match in matches:
            product_name = match["product_name"]
            family_name = match["family_name"]
            matched_keyword = match["matched_keyword"]
            source_field = match["source_field"]
            source_excerpt = match["source_excerpt"]
            match_confidence = match["match_confidence"]

            matched_families.add(family_name)
            product_row = product_name_to_ids.get(product_name)
            if not product_row:
                continue

            product_id, family_id = product_row

            cur.execute(
                """
                INSERT OR IGNORE INTO company_products
                (company_id, product_id, match_source, source_field, matched_keyword, match_confidence)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    product_id,
                    "profile text",
                    source_field,
                    matched_keyword,
                    match_confidence,
                ),
            )

            cur.execute(
                """
                INSERT INTO product_match_audit
                (company_id, product_id, family_id, source_field, matched_keyword, source_excerpt, match_confidence)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    product_id,
                    family_id,
                    source_field,
                    matched_keyword,
                    source_excerpt,
                    match_confidence,
                ),
            )

        for family_name in matched_families:
            family_id = family_name_to_id.get(family_name)
            if family_id:
                cur.execute(
                    """
                    INSERT OR IGNORE INTO product_family_matches (company_id, family_id)
                    VALUES (?, ?)
                    """,
                    (company_id, family_id),
                )

    sql_views = (
        "CREATE INDEX IF NOT EXISTS idx_companies_name ON companies(name);"
        "CREATE INDEX IF NOT EXISTS idx_companies_exporter_type ON companies(exporter_type);"
        "CREATE INDEX IF NOT EXISTS idx_companies_state ON companies(state);"
        "CREATE INDEX IF NOT EXISTS idx_company_products_company ON company_products(company_id);"
        "CREATE INDEX IF NOT EXISTS idx_contacts_company ON contacts(company_id);"

        "DROP VIEW IF EXISTS v_company_profile;"
        "CREATE VIEW v_company_profile AS "
        "SELECT "
        "c.id, c.name, c.exporter_type, c.website, c.address, c.state, c.postcode, c.countries_served, "
        "c.profile_url, "
        "GROUP_CONCAT(DISTINCT cert.name) AS certifications, "
        "GROUP_CONCAT(DISTINCT acc.name) AS accreditations "
        "FROM companies c "
        "LEFT JOIN company_certifications cc ON c.id = cc.company_id "
        "LEFT JOIN certifications cert ON cc.certification_id = cert.id "
        "LEFT JOIN company_accreditations ca ON c.id = ca.company_id "
        "LEFT JOIN accreditations acc ON ca.accreditation_id = acc.id "
        "GROUP BY c.id;"

        "DROP VIEW IF EXISTS v_company_products;"
        "CREATE VIEW v_company_products AS "
        "SELECT "
        "c.id, c.name, c.exporter_type, c.website, c.state, c.countries_served, c.profile_url, "
        "GROUP_CONCAT(DISTINCT pf.name) AS product_families, "
        "GROUP_CONCAT(DISTINCT CASE WHEN p.hierarchy_level = 2 THEN p.name END) AS product_variants "
        "FROM companies c "
        "LEFT JOIN company_products cp ON c.id = cp.company_id "
        "LEFT JOIN products p ON cp.product_id = p.id "
        "LEFT JOIN product_families pf ON p.family_id = pf.id "
        "GROUP BY c.id;"

        "DROP VIEW IF EXISTS v_product_hierarchy;"
        "CREATE VIEW v_product_hierarchy AS "
        "SELECT "
        "pf.name AS family_name, "
        "root.name AS root_product, "
        "child.name AS variant_product, "
        "child.hierarchy_level "
        "FROM product_families pf "
        "LEFT JOIN products root ON root.family_id = pf.id AND root.parent_product_id IS NULL "
        "LEFT JOIN products child ON child.parent_product_id = root.id "
        "ORDER BY pf.sort_order, child.name;"
    )

    cur.executescript(sql_views)

    conn.commit()
    conn.close()

# --------------------------------------------------
# EXPORT FROM DATABASE TO MASTER EXCEL
# --------------------------------------------------

def export_master_excel():
    conn = sqlite3.connect(DB_FILE)

    companies = pd.read_sql_query("SELECT * FROM companies ORDER BY name", conn)

    contacts = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            ct.email,
            ct.phone,
            ct.contact_source
        FROM contacts ct
        JOIN companies c ON c.id = ct.company_id
        ORDER BY c.name, ct.email, ct.phone
        """,
        conn,
    )

    certifications = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            cert.name AS certification
        FROM company_certifications cc
        JOIN companies c ON c.id = cc.company_id
        JOIN certifications cert ON cert.id = cc.certification_id
        ORDER BY c.name, cert.name
        """,
        conn,
    )

    accreditations = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            a.name AS accreditation
        FROM company_accreditations ca
        JOIN companies c ON c.id = ca.company_id
        JOIN accreditations a ON a.id = ca.accreditation_id
        ORDER BY c.name, a.name
        """,
        conn,
    )

    company_products = pd.read_sql_query("SELECT * FROM v_company_products ORDER BY name", conn)
    product_hierarchy = pd.read_sql_query("SELECT * FROM v_product_hierarchy", conn)
    company_profile = pd.read_sql_query("SELECT * FROM v_company_profile ORDER BY name", conn)

    # Build business-facing target dataset
    brokerage_targets = company_products.merge(
        company_profile[["id", "countries_served", "certifications", "accreditations"]],
        on="id",
        how="left",
    )

    contact_rollup = contacts.groupby("company_name", as_index=False).agg({
        "email": lambda s: "; ".join(unique_nonblank(s.tolist())),
        "phone": lambda s: "; ".join(unique_nonblank(s.tolist())),
    })

    brokerage_targets = brokerage_targets.merge(
        contact_rollup,
        left_on="name",
        right_on="company_name",
        how="left",
    ).drop(columns=["company_name"], errors="ignore")

    signals_df = brokerage_targets.apply(
        lambda r: pd.Series(compute_brokerage_signals(r)),
        axis=1,
    )
    brokerage_targets = pd.concat([brokerage_targets, signals_df], axis=1)
    brokerage_targets = brokerage_targets.sort_values(
        ["brokerage_fit_score", "name"],
        ascending=[False, True],
    )

    singapore_focus = brokerage_targets[
        brokerage_targets["singapore_fit"] == "Yes"
    ].copy()

    outreach_ready = brokerage_targets[
        brokerage_targets["outreach_ready"] == "Yes"
    ].copy()

    manual_verification = brokerage_targets[
        brokerage_targets["manual_checks_needed"].fillna("") != ""
    ].copy()

    product_match_audit = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            p.name AS product_name,
            pf.name AS product_family,
            a.source_field,
            a.matched_keyword,
            a.match_confidence,
            a.source_excerpt
        FROM product_match_audit a
        JOIN companies c ON c.id = a.company_id
        JOIN products p ON p.id = a.product_id
        JOIN product_families pf ON pf.id = a.family_id
        ORDER BY c.name, p.name, a.source_field
        """,
        conn,
    )

    product_summary = pd.read_sql_query(
        """
        SELECT
            pf.name AS product_family,
            COUNT(DISTINCT pfm.company_id) AS companies_with_product_family
        FROM product_family_matches pfm
        JOIN product_families pf ON pf.id = pfm.family_id
        GROUP BY pf.name
        ORDER BY companies_with_product_family DESC, pf.name
        """,
        conn,
    )

    certification_summary = pd.read_sql_query(
        """
        SELECT
            cert.name AS certification,
            COUNT(DISTINCT cc.company_id) AS companies_with_certification
        FROM company_certifications cc
        JOIN certifications cert ON cert.id = cc.certification_id
        GROUP BY cert.name
        ORDER BY companies_with_certification DESC, cert.name
        """,
        conn,
    )

    accreditation_summary = pd.read_sql_query(
        """
        SELECT
            a.name AS accreditation,
            COUNT(DISTINCT ca.company_id) AS companies_with_accreditation
        FROM company_accreditations ca
        JOIN accreditations a ON a.id = ca.accreditation_id
        GROUP BY a.name
        ORDER BY companies_with_accreditation DESC, a.name
        """,
        conn,
    )

    state_summary = pd.read_sql_query(
        """
        SELECT
            COALESCE(NULLIF(state, ''), '(blank)') AS state,
            COUNT(*) AS companies
        FROM companies
        GROUP BY COALESCE(NULLIF(state, ''), '(blank)')
        ORDER BY companies DESC, state
        """,
        conn,
    )

    exporter_type_summary = pd.read_sql_query(
        """
        SELECT
            COALESCE(NULLIF(exporter_type, ''), '(blank)') AS exporter_type,
            COUNT(*) AS companies
        FROM companies
        GROUP BY COALESCE(NULLIF(exporter_type, ''), '(blank)')
        ORDER BY companies DESC, exporter_type
        """,
        conn,
    )

    missing_data = pd.read_sql_query(
        """
        SELECT
            name,
            profile_url,
            website,
            exporter_type,
            state,
            countries_served,
            CASE WHEN COALESCE(website, '') = '' THEN 1 ELSE 0 END AS missing_website,
            CASE WHEN COALESCE(countries_served, '') = '' THEN 1 ELSE 0 END AS missing_countries,
            CASE WHEN NOT EXISTS (
                SELECT 1
                FROM contacts ct
                WHERE ct.company_id = companies.id
                  AND (COALESCE(ct.email, '') <> '' OR COALESCE(ct.phone, '') <> '')
            ) THEN 1 ELSE 0 END AS missing_contact,
            CASE WHEN NOT EXISTS (
                SELECT 1
                FROM company_products cp
                WHERE cp.company_id = companies.id
            ) THEN 1 ELSE 0 END AS missing_product_match
        FROM companies
        WHERE COALESCE(website, '') = ''
           OR COALESCE(countries_served, '') = ''
           OR NOT EXISTS (
                SELECT 1
                FROM contacts ct
                WHERE ct.company_id = companies.id
                  AND (COALESCE(ct.email, '') <> '' OR COALESCE(ct.phone, '') <> '')
           )
           OR NOT EXISTS (
                SELECT 1
                FROM company_products cp
                WHERE cp.company_id = companies.id
           )
        ORDER BY name
        """,
        conn,
    )

    summary = pd.DataFrame([
        {"Metric": "Generated UTC", "Value": datetime.now(UTC).isoformat(timespec="seconds")},
        {"Metric": "Companies", "Value": len(companies)},
        {"Metric": "Companies with Product Match", "Value": int(pd.read_sql_query("SELECT COUNT(DISTINCT company_id) AS n FROM company_products", conn).iloc[0, 0])},
        {"Metric": "Total Company-Product Links", "Value": int(pd.read_sql_query("SELECT COUNT(*) AS n FROM company_products", conn).iloc[0, 0])},
        {"Metric": "Companies with Certifications", "Value": int(pd.read_sql_query("SELECT COUNT(DISTINCT company_id) AS n FROM company_certifications", conn).iloc[0, 0])},
        {"Metric": "Total Company-Certification Links", "Value": int(pd.read_sql_query("SELECT COUNT(*) AS n FROM company_certifications", conn).iloc[0, 0])},
        {"Metric": "Companies with Accreditations", "Value": int(pd.read_sql_query("SELECT COUNT(DISTINCT company_id) AS n FROM company_accreditations", conn).iloc[0, 0])},
        {"Metric": "Total Company-Accreditation Links", "Value": int(pd.read_sql_query("SELECT COUNT(*) AS n FROM company_accreditations", conn).iloc[0, 0])},
        {"Metric": "Total Contact Rows", "Value": len(contacts)},
        {"Metric": "Singapore-fit Targets", "Value": int((brokerage_targets["singapore_fit"] == "Yes").sum())},
        {"Metric": "Outreach-ready Targets", "Value": int((brokerage_targets["outreach_ready"] == "Yes").sum())},
        {"Metric": "Companies Missing Website", "Value": int((companies["website"].fillna("") == "").sum())},
        {"Metric": "Companies Missing Countries Served", "Value": int((companies["countries_served"].fillna("") == "").sum())},
    ])

    conn.close()

    companies_for_excel = companies.copy().drop(columns=["details_json"], errors="ignore")

    if "page_text_excerpt" in companies_for_excel.columns:
        companies_for_excel["page_text_preview"] = (
            companies_for_excel["page_text_excerpt"]
            .fillna("")
            .astype(str)
            .str[:220]
        )
        companies_for_excel = companies_for_excel.drop(columns=["page_text_excerpt"], errors="ignore")

    final_excel_path, used_fallback = safe_write_excel(
        EXCEL_FILE,
        {
            "Summary": summary,
            "Brokerage Targets": brokerage_targets,
            "Singapore Focus": singapore_focus,
            "Outreach Ready": outreach_ready,
            "Manual Verification": manual_verification,
            "Product Summary": product_summary,
            "Certification Summary": certification_summary,
            "Accreditation Summary": accreditation_summary,
            "State Summary": state_summary,
            "Exporter Type Summary": exporter_type_summary,
            "Missing Data": missing_data,
            "Company Profile": company_profile,
            "Company Products": company_products,
            "Companies": companies_for_excel,
            "Contacts": contacts,
            "Product Match Audit": product_match_audit,
            "Certifications": certifications,
            "Accreditations": accreditations,
            "Product Hierarchy": product_hierarchy,
        },
    )

    dashboard_polish_workbook(final_excel_path)
    return final_excel_path, used_fallback

# --------------------------------------------------
# SCHEMA DOCS / RUNNER
# --------------------------------------------------

def export_schema_files():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        "SELECT type, name, sql FROM sqlite_master "
        "WHERE sql IS NOT NULL AND name != 'sqlite_sequence' "
        "ORDER BY type, name"
    )
    rows = cur.fetchall()

    sql_lines = [row[2] + ";" for row in rows if row[2]]
    Path(SCHEMA_SQL_FILE).write_text("\n\n".join(sql_lines), encoding="utf-8")

    md_parts = [
        "# Exporters Database Schema",
        "",
        f"Generated: {datetime.now(UTC).isoformat(timespec='seconds')}",
        "",
        "## Overview",
        "This project stores exporter listings, contacts, certifications, accreditations, products, countries served, and reporting views in exporters_final.db.",
        "",
        "## Main Reporting Views",
        "- **v_company_profile**: one row per company with certifications, accreditations, and countries served rolled up.",
        "- **v_company_products**: one row per company with product family and product variant rollups.",
        "- **v_product_hierarchy**: product family hierarchy reference.",
        "",
        "## Known Limitations",
        "- Product matching is keyword based, not ML-based.",
        "- Contacts are extracted from visible text and links, so some false positives/negatives can still occur.",
        "- Some exporter pages contain repeated site boilerplate which is filtered on a best-effort basis.",
        "- Countries serviced are extracted only when surfaced in the public profile content.",
        "",
    ]

    purpose_map = {
        "companies": "Master exporter/company records.",
        "contacts": "Extracted contact details by company.",
        "attributes": "Raw extracted text attributes kept for traceability.",
        "certifications": "Certification lookup table.",
        "company_certifications": "Many-to-many link between companies and certifications.",
        "accreditations": "Accreditation lookup table.",
        "company_accreditations": "Many-to-many link between companies and accreditations.",
        "product_families": "Top-level product family dimension.",
        "products": "Product dimension with hierarchy metadata.",
        "company_products": "Matched company-to-product relationships.",
        "product_family_matches": "Matched company-to-product-family relationships.",
        "product_match_audit": "Audit table showing why product matches were made.",
        "v_company_profile": "Profile reporting view.",
        "v_company_products": "Company-product reporting view.",
        "v_product_hierarchy": "Hierarchy reference view.",
    }

    for obj_type, name, sql in rows:
        md_parts.append(f"## {obj_type.title()}: {name}")
        md_parts.append("")
        if name in purpose_map:
            md_parts.append(purpose_map[name])
            md_parts.append("")
        md_parts.append("```sql")
        md_parts.append(sql + ";")
        md_parts.append("```")
        md_parts.append("")

    Path(SCHEMA_MD_FILE).write_text("\n".join(md_parts), encoding="utf-8")
    conn.close()


def write_run_bat():
    content = (
        "@echo off\n"
        "cd /d \"%~dp0\"\n"
        "python -m py_compile pipeline_final.py\n"
        "if errorlevel 1 goto :end\n"
        "python -u pipeline_final.py\n"
        ":end\n"
        "pause\n"
    )
    Path(RUN_BAT_FILE).write_text(content, encoding="utf-8")

# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main():
    print("Collecting listings...")
    listings = collect_listings()
    print(f"Collected {len(listings)} rows")

    print("Collecting profile details...")
    enriched = enrich_rows(listings)
    df = clean_dataframe(pd.DataFrame(enriched))

    print("Building database...")
    build_database(df)

    print("Exporting Excel from database...")
    final_excel_path, used_fallback = export_master_excel()

    export_schema_files()
    write_run_bat()
    cleanup_fallback_excels_if_master_written(used_fallback)

    print("FINAL PIPELINE COMPLETE")
    print("Created:")
    print(f"- {DB_FILE}")
    print(f"- {final_excel_path}")
    print(f"- {SCHEMA_SQL_FILE}")
    print(f"- {SCHEMA_MD_FILE}")
    print(f"- {RUN_BAT_FILE}")

    if used_fallback:
        print("exporters_master.xlsx was locked (likely open in Excel).")
        print("A fallback Excel file was created.")
        print("It will automatically replace exporters_master.xlsx after Excel is closed.")
    else:
        print("Database, Excel, schema files, and batch runner are fully in sync.")


if __name__ == "__main__":
    if len(sys.argv) >= 4 and sys.argv[1] == "--finalize-fallback":
        finalize_locked_excel(sys.argv[2], sys.argv[3])
        sys.exit(0)

    main()