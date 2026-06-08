import os
import sys
import re
import json
import time
import html
import sqlite3
import subprocess
from datetime import datetime, UTC
from pathlib import Path
from urllib.parse import urljoin
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# --------------------------------------------------
# CONFIG
# --------------------------------------------------

BASE_URL = "https://www.aussiemeattradehub.com.au"
LIST_URL = BASE_URL + "/RMED/search/GetListingByPagination"
SEARCH_PAGE_URL = BASE_URL + "/rmed/search"

DB_FILE = "exporters_final.db"
EXCEL_FILE = "exporters_master.xlsx"
SCHEMA_SQL_FILE = "exporters_final_schema.sql"
SCHEMA_MD_FILE = "exporters_schema.md"
RUN_BAT_FILE = "run_pipeline.bat"

MAX_PAGES = 100
REQUEST_DELAY_SECONDS = 0.25
REQUEST_TIMEOUT = 60
MAX_WORKERS = 8

HEADERS = {
    "accept": "application/json, text/javascript, */*; q=0.01",
    "content-type": "application/json; charset=UTF-8",
    "origin": BASE_URL,
    "referer": SEARCH_PAGE_URL,
    "x-requested-with": "XMLHttpRequest",
    "user-agent": "Mozilla/5.0",
}

SEARCH_PAYLOAD = {
    "search": {
        "SearchTerm": "",
        "SearchCategory": "All",
    }
}

FIELD_WEIGHTS = {
    "description": 1.00,
    "page_heading": 0.90,
    "meta_description": 0.70,
    "details_json": 0.60,
    "page_text_excerpt": 0.50,
}

KNOWN_CERTIFICATIONS = [
    "Certified Pasturefed",
    "Grainfed",
    "Halal",
    "MSA",
    "Organic",
    "Other 3rd Party Audited Verified",
]

KNOWN_ACCREDITATIONS = [
    "China",
    "EU",
    "Malaysia",
    "Egypt",
    "Indonesia",
    "Russia",
]

PRODUCT_HIERARCHY = {
    "Beef": {
        "Beef": ["beef"],
        "Grainfed Beef": ["grainfed beef", "grain-fed beef"],
        "Grassfed Beef": ["grassfed beef", "grass-fed beef"],
        "Wagyu Beef": ["wagyu beef", "wagyu"],
        "Beef Offal": ["beef offal"],
        "Value Added Beef": [
            "value added beef",
            "value-added beef",
            "value added beef products",
            "value-added beef products",
        ],
        "Veal": ["veal"],
    },
    "Lamb and Sheepmeat": {
        "Lamb": ["lamb"],
        "Mutton": ["mutton"],
        "Sheepmeat": ["sheepmeat"],
        "Sheep Offal": ["sheep offal"],
        "Value Added Lamb/Mutton": [
            "value added lamb",
            "value added lamb/mutton",
            "value-added lamb",
            "value-added lamb/mutton",
            "value added lamb/mutton products",
            "value-added lamb/mutton products",
        ],
    },
    "Goat": {
        "Goatmeat": ["goatmeat", "goat meat"],
        "Goat Offal": ["goat offal"],
        "Value Added Goatmeat": [
            "value added goatmeat",
            "value-added goatmeat",
            "value added goatmeat products",
            "value-added goatmeat products",
        ],
    },
}

BOILERPLATE_PATTERNS = [
    r"ListingDetails",
    r"Exporters Database",
    r"Brand(?:,| &| &amp;) Licensing(?: & Assets| &amp; Assets)?",
    r"About the brand",
    r"Licensing program",
    r"Manage my licence",
    r"Assets",
    r"Global Insights",
    r"Trade Shows",
    r"Welcome,?",
    r"Report Centre",
    r"Login",
    r"Signup",
    r"Error Enquiry unable to send\.? Please try after sometime",
    r"Ok Processing This enquiry is being sent to the exporter - please wait",
    r"Successful Your enquiry was successfully(?: sent)?",
    r"×",
]

ABN_NEARBY_PATTERN = re.compile(r"\bABN\b", re.IGNORECASE)
EMAIL_PATTERN = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_PATTERN = re.compile(r"(?:\+?\d[\d\s\-()]{7,}\d)")
YEAR_RANGE_PATTERN = re.compile(r"^\d{4}\s*[-–]\s*\d{4}$")


# --------------------------------------------------
# GENERAL HELPERS
# --------------------------------------------------

def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def strip_boilerplate(text):
    text = clean_text(text)
    for pattern in BOILERPLATE_PATTERNS:
        text = re.sub(pattern, " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def extract_between(text, start_label, end_labels):
    text = clean_text(text)
    start_match = re.search(re.escape(start_label) + r"\s+(.*)", text, flags=re.IGNORECASE)
    if not start_match:
        return ""

    remainder = start_match.group(1)
    end_positions = []

    for label in end_labels:
        m = re.search(r"\s" + re.escape(label) + r"\b", remainder, flags=re.IGNORECASE)
        if m:
            end_positions.append(m.start())

    if end_positions:
        return clean_text(remainder[:min(end_positions)])

    return clean_text(remainder)


def flatten_hierarchy_to_product_map():
    out = {}
    for family, products in PRODUCT_HIERARCHY.items():
        for product_name, keywords in products.items():
            out[product_name] = {"family": family, "keywords": keywords}
    return out


def unique_nonblank(values):
    seen = set()
    out = []
    for value in values:
        value = clean_text(value)
        if value and value not in seen:
            out.append(value)
            seen.add(value)
    return out


def normalize_phone(phone):
    phone = clean_text(phone)
    phone = re.sub(r"[^\d+()\-\ ]", "", phone)
    phone = re.sub(r"\s+", " ", phone).strip()
    return phone


def is_probable_abn(candidate, surrounding_text=""):
    digits = re.sub(r"\D", "", candidate)
    return len(digits) == 11 and bool(ABN_NEARBY_PATTERN.search(surrounding_text or ""))


def filter_phone_candidates(candidates, surrounding_text=""):
    cleaned = []
    for raw in candidates:
        candidate = normalize_phone(raw)
        digits = re.sub(r"\D", "", candidate)
        if not candidate:
            continue
        if YEAR_RANGE_PATTERN.match(candidate):
            continue
        if len(digits) < 8:
            continue
        if is_probable_abn(candidate, surrounding_text):
            continue
        cleaned.append(candidate)
    return unique_nonblank(cleaned)


def build_field_texts_for_matching(company_row):
    return {
        "description": clean_text(company_row.get("description", "")),
        "page_heading": clean_text(company_row.get("page_heading", "")),
        "meta_description": clean_text(company_row.get("meta_description", "")),
        "details_json": clean_text(company_row.get("details_json", "")),
        "page_text_excerpt": clean_text(company_row.get("page_text_excerpt", "")),
    }


def score_product_matches(field_texts, product_map):
    matches = []

    for source_field, source_text in field_texts.items():
        lower_text = source_text.lower()
        if not lower_text:
            continue

        for product_name, meta in product_map.items():
            for keyword in meta["keywords"]:
                if keyword.lower() in lower_text:
                    confidence = FIELD_WEIGHTS.get(source_field, 0.50)
                    matches.append({
                        "product_name": product_name,
                        "family_name": meta["family"],
                        "matched_keyword": keyword,
                        "source_field": source_field,
                        "source_excerpt": source_text[:250],
                        "match_confidence": confidence,
                    })
                    break

    return matches


# --------------------------------------------------
# EXCEL WRITING / DASHBOARD STYLING
# --------------------------------------------------

def write_styled_excel(path, sheets):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

        for url_col_name in ["Profile URL", "Image URL", "Website"]:
            col_idx = headers.get(url_col_name)
            if col_idx:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        cell.hyperlink = str(cell.value)
                        cell.style = "Hyperlink"

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)

    wb.save(path)


def dashboard_polish_workbook(path):
    wb = load_workbook(path)

    desired_order = [
        "Summary",
        "Product Summary",
        "Certification Summary",
        "Accreditation Summary",
        "State Summary",
        "Exporter Type Summary",
        "Missing Data",
        "Company Profile",
        "Company Products",
        "Companies",
        "Contacts",
        "Product Match Audit",
        "Certifications",
        "Accreditations",
        "Product Hierarchy",
    ]
    ordered = [wb[s] for s in desired_order if s in wb.sheetnames] + [wb[s] for s in wb.sheetnames if s not in desired_order]
    wb._sheets = ordered

    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    blue_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    pink_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws.sheet_view.showGridLines = False
        for row in range(2, ws.max_row + 1):
            metric = ws.cell(row=row, column=1).value
            val_cell = ws.cell(row=row, column=2)
            if metric and metric != "Generated UTC":
                val_cell.fill = green_fill
                val_cell.font = Font(bold=True)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 22

    for sheet_name in [
        "Product Summary",
        "Certification Summary",
        "Accreditation Summary",
        "State Summary",
        "Exporter Type Summary",
    ]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.sheet_view.showGridLines = False
            if ws.max_row >= 2 and ws.max_column >= 2:
                rng = f"B2:B{ws.max_row}"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="63BE7B",
                    ),
                )
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=2).font = Font(bold=True)

    for sheet_name in ["Company Profile", "Company Products", "Companies"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = {
                str(ws.cell(row=1, column=i).value or "").strip().lower(): i
                for i in range(1, ws.max_column + 1)
            }

            for key in [
                "name",
                "company_name",
                "exporter_type",
                "product_families",
                "product_variants",
                "certifications",
                "accreditations",
                "state",
            ]:
                idx = headers.get(key)
                if idx:
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=idx).fill = blue_fill

            for text_col in ["page_text_preview", "meta_description", "description"]:
                idx = headers.get(text_col)
                if idx:
                    ws.column_dimensions[get_column_letter(idx)].width = 40

    if "Contacts" in wb.sheetnames:
        ws = wb["Contacts"]
        ws.sheet_view.showGridLines = False
        for row in range(2, ws.max_row + 1):
            email = str(ws.cell(row=row, column=2).value or "").strip()
            phone = str(ws.cell(row=row, column=3).value or "").strip()
            if email or phone:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill

    if "Missing Data" in wb.sheetnames:
        ws = wb["Missing Data"]
        ws.sheet_view.showGridLines = False
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = pink_fill

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.topLeftCell = "A1"

    wb.save(path)


# --------------------------------------------------
# EXCEL LOCK / FALLBACK HANDLING
# --------------------------------------------------

def cleanup_fallback_excels_if_master_written(used_fallback):
    if used_fallback:
        print("⚠️ Keeping fallback Excel files (main file was locked).")
        return

    files = list(Path(".").glob("exporters_master_locked_export_*.xlsx"))
    if not files:
        return

    for f in files:
        try:
            f.unlink()
            print(f"Deleted fallback file: {f.name}")
        except Exception as e:
            print(f"Could not delete {f.name}: {e}")

    print("All fallback Excel files cleaned up.")


def finalize_locked_excel(master_path, fallback_path, poll_seconds=3, max_attempts=1200):
    master_path = os.path.abspath(master_path)
    fallback_path = os.path.abspath(fallback_path)

    for _ in range(max_attempts):
        if not os.path.exists(fallback_path):
            return
        try:
            os.replace(fallback_path, master_path)
            for f in Path(".").glob("exporters_master_locked_export_*.xlsx"):
                try:
                    if os.path.abspath(str(f)) not in {master_path, fallback_path}:
                        f.unlink()
                except Exception:
                    pass
            return
        except PermissionError:
            time.sleep(poll_seconds)
        except Exception:
            time.sleep(poll_seconds)


def launch_finalize_in_background(master_path, fallback_path):
    python_exe = sys.executable
    script_path = os.path.abspath(__file__)

    creation_flags = 0
    if os.name == "nt":
        creation_flags = (
            getattr(subprocess, "DETACHED_PROCESS", 0)
            | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
        )

    subprocess.Popen(
        [python_exe, script_path, "--finalize-fallback", master_path, fallback_path],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        stdin=subprocess.DEVNULL,
        close_fds=True,
        creationflags=creation_flags,
    )


def safe_write_excel(path, sheets):
    try:
        write_styled_excel(path, sheets)
        return path, False
    except PermissionError:
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        fallback_path = path.replace(".xlsx", f"_locked_export_{timestamp}.xlsx")
        write_styled_excel(fallback_path, sheets)
        launch_finalize_in_background(path, fallback_path)
        return fallback_path, True


# --------------------------------------------------
# SCRAPING LISTINGS
# --------------------------------------------------

def fetch_list_page(page_number):
    payload = dict(SEARCH_PAYLOAD)
    payload["PageNumber"] = page_number

    response = requests.post(LIST_URL, headers=HEADERS, json=payload, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()

    try:
        data = response.json()
    except ValueError:
        data = response.text

    if isinstance(data, dict):
        html_snippet = data.get("Html") or data.get("html") or ""
    elif isinstance(data, str):
        html_snippet = data
    else:
        html_snippet = ""

    return html.unescape(html_snippet)


def parse_listing_html(html_snippet):
    soup = BeautifulSoup(html_snippet, "html.parser")
    rows = []

    for c in soup.select("div.company_list"):
        header_link = c.select_one("div.company_list_header a")
        desc_tag = c.select_one("div.company_list_body p")
        img_tag = c.select_one("img")

        rows.append({
            "company_name": header_link.get_text(strip=True) if header_link else "",
            "description": strip_boilerplate(desc_tag.get_text(" ", strip=True) if desc_tag else ""),
            "profile_url": urljoin(BASE_URL, header_link["href"]) if header_link and header_link.has_attr("href") else "",
            "image_url": urljoin(BASE_URL, img_tag["src"]) if img_tag and img_tag.has_attr("src") else "",
        })

    return rows


def collect_listings(max_pages=MAX_PAGES):
    all_rows = []

    for page in range(1, max_pages + 1):
        rows = parse_listing_html(fetch_list_page(page))
        print(f"Page {page}: {len(rows)}")

        if not rows:
            break

        all_rows.extend(rows)
        time.sleep(REQUEST_DELAY_SECONDS)

    return all_rows


# --------------------------------------------------
# SCRAPING PROFILES
# --------------------------------------------------
def safe_get(url, timeout=20):
    try:
        response = requests.get(
            url,
            headers={"user-agent": HEADERS["user-agent"]},
            timeout=timeout,
        )
        response.raise_for_status()
        return response
    except Exception:
        return None


def extract_website_from_details(details, page_text=""):
    # First try structured fields
    for k, v in details.items():
        key = clean_text(k).lower()
        val = clean_text(v)

        if "website" in key and val:
            if not val.startswith(("http://", "https://")):
                val = "https://" + val
            return val

    # Fallback: look in page text
    m = re.search(r"(https?://\S+|www\.\S+)", page_text, flags=re.IGNORECASE)
    if m:
        website = clean_text(m.group(1))
        if not website.startswith(("http://", "https://")):
            website = "https://" + website
        return website

    return ""


def extract_emails_from_website(url):
    if not url:
        return []

    emails = []

    resp = safe_get(url)
    if not resp:
        return []

    html_text = resp.text
    emails.extend(EMAIL_PATTERN.findall(html_text))

    soup = BeautifulSoup(html_text, "html.parser")

    # Look for obvious contact page links
    contact_links = []
    for a in soup.select("a[href]"):
        href = clean_text(a.get("href", ""))
        text = clean_text(a.get_text(" ", strip=True)).lower()

        if not href:
            continue

        href_lower = href.lower()
        if "contact" in href_lower or "contact" in text:
            contact_links.append(urljoin(url, href))

    # Check a small number of contact pages only
    for link in unique_nonblank(contact_links)[:2]:
        resp2 = safe_get(link)
        if resp2:
            emails.extend(EMAIL_PATTERN.findall(resp2.text))

    return unique_nonblank(emails)

def parse_profile_page(html_text):
    soup = BeautifulSoup(html_text, "html.parser")
    page_text_raw = soup.get_text(" ", strip=True)
    page_text = strip_boilerplate(page_text_raw)

    def find_meta(names):
        for name in names:
            el = soup.select_one(f'meta[property="{name}"]') or soup.select_one(f'meta[name="{name}"]')
            if el and el.has_attr("content"):
                return strip_boilerplate(el["content"])
        return ""

    # ------------------------------------------
    # CONTACTS FROM PROFILE PAGE
    # ------------------------------------------

    href_contacts = [
        a.get("href", "")
        for a in soup.select('a[href^="mailto:"], a[href^="tel:"]')
    ]
    href_emails = [
        x.replace("mailto:", "").strip()
        for x in href_contacts
        if x.startswith("mailto:")
    ]
    href_phones = [
        x.replace("tel:", "").strip()
        for x in href_contacts
        if x.startswith("tel:")
    ]

    regex_emails = EMAIL_PATTERN.findall(page_text)
    regex_phones = PHONE_PATTERN.findall(page_text)

    emails = unique_nonblank(href_emails + regex_emails)
    phones = filter_phone_candidates(
        href_phones + regex_phones,
        surrounding_text=page_text
    )

    # ------------------------------------------
    # STRUCTURED DETAILS
    # ------------------------------------------

    details = {}

    for dt in soup.select("dt"):
        dd = dt.find_next_sibling("dd")
        if dd:
            k = clean_text(dt.get_text(" ", strip=True))
            v = strip_boilerplate(dd.get_text(" ", strip=True))
            if k and v:
                details[k] = v

    for row in soup.select("table tr"):
        cells = row.find_all(["th", "td"])
        if len(cells) == 2:
            k = clean_text(cells[0].get_text(" ", strip=True))
            v = strip_boilerplate(cells[1].get_text(" ", strip=True))
            if k and v and k not in details:
                details[k] = v

    # ------------------------------------------
    # WEBSITE EXTRACTION
    # ------------------------------------------

    website = extract_website_from_details(details, page_text)

    # ------------------------------------------
    # EXTRA EMAIL EXTRACTION
    # ------------------------------------------

    # from structured fields
    for v in details.values():
        emails.extend(EMAIL_PATTERN.findall(v))

    # from raw HTML
    emails.extend(EMAIL_PATTERN.findall(html_text))

    # from company website / contact page(s)
    website_emails = extract_emails_from_website(website)
    emails.extend(website_emails)

    # final dedupe
    emails = unique_nonblank(emails)

    details_json = json.dumps(details, ensure_ascii=False) if details else ""

    return {
        "page_title": strip_boilerplate(soup.title.get_text(strip=True) if soup.title else ""),
        "page_heading": strip_boilerplate(
            soup.select_one("h1").get_text(" ", strip=True)
            if soup.select_one("h1") else ""
        ),
        "meta_description": find_meta(["description", "og:description"]),
        "meta_title": find_meta(["og:title"]),
        "website": website,
        "emails": "; ".join(emails),
        "phones": "; ".join(phones),
        "details_json": details_json,
        "page_text_excerpt": page_text[:4000],
    }


def fetch_profile_details(profile_url):
    if not profile_url:
        return {}
    response = requests.get(profile_url, headers={"user-agent": HEADERS["user-agent"]}, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    return parse_profile_page(response.text)


def enrich_one_row(row):
    profile_url = row.get("profile_url", "")
    try:
        details = fetch_profile_details(profile_url) if profile_url else {}
    except Exception as e:
        details = {"profile_error": str(e)}

    merged = dict(row)
    merged.update(details)
    return merged


def enrich_rows(rows):
    enriched = []
    total = len(rows)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(enrich_one_row, row) for row in rows]

        completed = 0
        for future in as_completed(futures):
            completed += 1
            print(f"Profile {completed}/{total}")
            enriched.append(future.result())

    enriched = sorted(
        enriched,
        key=lambda r: (
            clean_text(r.get("company_name", "")),
            clean_text(r.get("profile_url", "")),
        ),
    )
    return enriched


def clean_dataframe(df):
    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]):
            df[col] = df[col].fillna("").astype(str).map(clean_text)

    df.insert(0, "exported_at_utc", datetime.now(UTC).isoformat(timespec="seconds"))

    dedupe_cols = [c for c in ["company_name", "profile_url"] if c in df.columns]
    if dedupe_cols:
        df = df.drop_duplicates(subset=dedupe_cols, keep="first").reset_index(drop=True)

    if "company_name" in df.columns:
        df = df.sort_values(by=["company_name", "profile_url"], kind="stable").reset_index(drop=True)

    return df


# --------------------------------------------------
# DATABASE SETUP
# --------------------------------------------------

def init_schema(cur):
    cur.execute("PRAGMA foreign_keys = OFF;")

    sql = (
        "DROP VIEW IF EXISTS v_product_hierarchy;"
        "DROP VIEW IF EXISTS v_company_products;"
        "DROP VIEW IF EXISTS v_company_profile;"
        "DROP TABLE IF EXISTS product_match_audit;"
        "DROP TABLE IF EXISTS company_certifications;"
        "DROP TABLE IF EXISTS certifications;"
        "DROP TABLE IF EXISTS company_accreditations;"
        "DROP TABLE IF EXISTS accreditations;"
        "DROP TABLE IF EXISTS product_families;"
        "DROP TABLE IF EXISTS products;"
        "DROP TABLE IF EXISTS company_products;"
        "DROP TABLE IF EXISTS product_family_matches;"
        "DROP TABLE IF EXISTS contacts;"
        "DROP TABLE IF EXISTS attributes;"
        "DROP TABLE IF EXISTS companies;"

        "CREATE TABLE companies ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "name TEXT, description TEXT, profile_url TEXT UNIQUE, image_url TEXT,"
        "exporter_type TEXT, licence_number TEXT, establishment_numbers TEXT,"
        "website TEXT, address TEXT, abn TEXT, state TEXT, postcode TEXT,"
        "page_title TEXT, page_heading TEXT, meta_description TEXT, meta_title TEXT,"
        "page_text_excerpt TEXT, details_json TEXT, profile_error TEXT);"

        "CREATE TABLE contacts ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, email TEXT, phone TEXT, contact_source TEXT,"
        "FOREIGN KEY(company_id) REFERENCES companies(id));"

        "CREATE TABLE attributes ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, name TEXT, value TEXT,"
        "FOREIGN KEY(company_id) REFERENCES companies(id));"

        "CREATE TABLE certifications (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE);"
        "CREATE TABLE company_certifications (company_id INTEGER, certification_id INTEGER, UNIQUE(company_id, certification_id), FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(certification_id) REFERENCES certifications(id));"

        "CREATE TABLE accreditations (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE);"
        "CREATE TABLE company_accreditations (company_id INTEGER, accreditation_id INTEGER, UNIQUE(company_id, accreditation_id), FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(accreditation_id) REFERENCES accreditations(id));"

        "CREATE TABLE product_families (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, sort_order INTEGER);"
        "CREATE TABLE products (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE, family_id INTEGER, parent_product_id INTEGER, hierarchy_level INTEGER, FOREIGN KEY(family_id) REFERENCES product_families(id), FOREIGN KEY(parent_product_id) REFERENCES products(id));"

        "CREATE TABLE company_products ("
        "company_id INTEGER, product_id INTEGER, match_source TEXT, source_field TEXT, matched_keyword TEXT, match_confidence REAL,"
        "UNIQUE(company_id, product_id, matched_keyword, source_field),"
        "FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(product_id) REFERENCES products(id));"

        "CREATE TABLE product_family_matches (company_id INTEGER, family_id INTEGER, UNIQUE(company_id, family_id), FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(family_id) REFERENCES product_families(id));"

        "CREATE TABLE product_match_audit ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, product_id INTEGER, family_id INTEGER, source_field TEXT, matched_keyword TEXT, source_excerpt TEXT, match_confidence REAL,"
        "FOREIGN KEY(company_id) REFERENCES companies(id), FOREIGN KEY(product_id) REFERENCES products(id), FOREIGN KEY(family_id) REFERENCES product_families(id));"
    )

    cur.executescript(sql)
    cur.execute("PRAGMA foreign_keys = ON;")


def upsert_lookup_and_link(cur, company_id, table, link_table, link_fk, value):
    cur.execute(f"INSERT OR IGNORE INTO {table} (name) VALUES (?)", (value,))
    cur.execute(f"SELECT id FROM {table} WHERE name = ?", (value,))
    item_id = cur.fetchone()[0]
    cur.execute(f"INSERT OR IGNORE INTO {link_table} (company_id, {link_fk}) VALUES (?, ?)", (company_id, item_id))


def parse_company_fields(text):
    fields = {
        "website": "",
        "address": "",
        "abn": "",
        "exporter_type": "",
        "licence_number": "",
        "establishment_numbers": "",
        "state": "",
        "postcode": "",
    }

    m = re.search(r"Website\s+(https?://\S+|www\.\S+)", text, flags=re.IGNORECASE)
    if m:
        fields["website"] = clean_text(m.group(1))

    fields["address"] = extract_between(
        text,
        "Address",
        ["ABN", "Exporter Type", "Exporter type", "Licence number", "Establishment numbers", "Accreditations", "Certifications", "Login"],
    )

    m = re.search(r"ABN\s+([A-Z0-9 ]+?)(?=\s+(Exporter Type|Exporter type|Licence number|Establishment numbers|Accreditations|Certifications|Login)\b)", text, flags=re.IGNORECASE)
    if m:
        fields["abn"] = clean_text(m.group(1))

    m = re.search(r"Exporter\s+[Tt]ype\s+(.*?)(?=\s+(Licence number|Establishment numbers|Accreditations|Certifications|Login)\b)", text, flags=re.IGNORECASE)
    if m:
        fields["exporter_type"] = clean_text(m.group(1))

    m = re.search(r"Licence number\s+(.*?)(?=\s+(Establishment numbers|Accreditations|Certifications|Login)\b)", text, flags=re.IGNORECASE)
    if m:
        fields["licence_number"] = clean_text(m.group(1))

    m = re.search(r"Establishment numbers\s+(.*?)(?=\s+(Accreditations|Certifications|Login)\b)", text, flags=re.IGNORECASE)
    if m:
        fields["establishment_numbers"] = clean_text(m.group(1))

    m = re.search(r"\b(NSW|VIC|QLD|WA|SA|TAS|ACT|NT)\s+(\d{4})\b", fields["address"])
    if m:
        fields["state"] = m.group(1)
        fields["postcode"] = m.group(2)

    return fields


def build_database(df):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    init_schema(cur)

    for _, row in df.iterrows():
        cur.execute(
            "INSERT OR IGNORE INTO companies (name, description, profile_url, image_url, page_title, page_heading, meta_description, meta_title, page_text_excerpt, details_json, profile_error) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                clean_text(row.get("company_name")),
                clean_text(row.get("description")),
                clean_text(row.get("profile_url")),
                clean_text(row.get("image_url")),
                clean_text(row.get("page_title")),
                clean_text(row.get("page_heading")),
                clean_text(row.get("meta_description")),
                clean_text(row.get("meta_title")),
                clean_text(row.get("page_text_excerpt")),
                clean_text(row.get("details_json")),
                clean_text(row.get("profile_error")),
            ),
        )

        cur.execute("SELECT id FROM companies WHERE profile_url = ?", (clean_text(row.get("profile_url")),))
        company_id = cur.fetchone()[0]

        email_values = unique_nonblank(clean_text(row.get("emails")).split(";")) if clean_text(row.get("emails")) else []
        phone_values = unique_nonblank(clean_text(row.get("phones")).split(";")) if clean_text(row.get("phones")) else []

        for email in email_values:
            if email and "mla.com.au" not in email.lower():
                cur.execute(
                    "INSERT INTO contacts (company_id, email, phone, contact_source) VALUES (?, ?, ?, ?)",
                    (company_id, email, "", "regex_or_href"),
                )

        for phone in phone_values:
            if phone:
                cur.execute(
                    "INSERT INTO contacts (company_id, email, phone, contact_source) VALUES (?, ?, ?, ?)",
                    (company_id, "", phone, "regex_or_href"),
                )

        for raw_name in ["page_title", "page_heading", "meta_description", "meta_title", "page_text_excerpt", "details_json"]:
            value = clean_text(row.get(raw_name))
            if value:
                cur.execute(
                    "INSERT INTO attributes (company_id, name, value) VALUES (?, ?, ?)",
                    (company_id, raw_name, value),
                )

        text = " ".join([
            clean_text(row.get("description")),
            clean_text(row.get("page_heading")),
            clean_text(row.get("page_text_excerpt")),
            clean_text(row.get("details_json")),
        ])

        fields = parse_company_fields(text)
        cur.execute(
            "UPDATE companies SET website = ?, address = ?, abn = ?, exporter_type = ?, licence_number = ?, establishment_numbers = ?, state = ?, postcode = ? WHERE id = ?",
            (
                fields["website"],
                fields["address"],
                fields["abn"],
                fields["exporter_type"],
                fields["licence_number"],
                fields["establishment_numbers"],
                fields["state"],
                fields["postcode"],
                company_id,
            ),
        )

        cert_text = extract_between(text, "Certifications", ["Login", "Australia", "Privacy", "View Report"])
        for cert in [c for c in KNOWN_CERTIFICATIONS if c in cert_text]:
            upsert_lookup_and_link(cur, company_id, "certifications", "company_certifications", "certification_id", cert)

        acc_text = extract_between(text, "Accreditations", ["Certifications", "Login", "Australia", "Privacy", "View Report"])
        for acc in [a for a in KNOWN_ACCREDITATIONS if re.search(r"\b" + re.escape(a) + r"\b", acc_text)]:
            upsert_lookup_and_link(cur, company_id, "accreditations", "company_accreditations", "accreditation_id", acc)

    family_order = list(PRODUCT_HIERARCHY.keys())
    for idx, family_name in enumerate(family_order, start=1):
        cur.execute(
            "INSERT OR IGNORE INTO product_families (name, sort_order) VALUES (?, ?)",
            (family_name, idx),
        )

    for family_name, variants in PRODUCT_HIERARCHY.items():
        cur.execute("SELECT id FROM product_families WHERE name = ?", (family_name,))
        family_id = cur.fetchone()[0]

        cur.execute(
            "INSERT OR IGNORE INTO products (name, family_id, parent_product_id, hierarchy_level) VALUES (?, ?, NULL, 1)",
            (family_name, family_id),
        )
        cur.execute("SELECT id FROM products WHERE name = ?", (family_name,))
        root_product_id = cur.fetchone()[0]

        for product_name in variants.keys():
            if product_name == family_name:
                continue
            cur.execute(
                "INSERT OR IGNORE INTO products (name, family_id, parent_product_id, hierarchy_level) VALUES (?, ?, ?, 2)",
                (product_name, family_id, root_product_id),
            )

    product_map = flatten_hierarchy_to_product_map()

    cur.execute("""
        SELECT
            id,
            description,
            page_heading,
            meta_description,
            details_json,
            page_text_excerpt
        FROM companies
    """)
    company_rows = cur.fetchall()

    for company_id, description, page_heading, meta_description, details_json, page_text_excerpt in company_rows:
        company_row = {
            "description": description,
            "page_heading": page_heading,
            "meta_description": meta_description,
            "details_json": details_json,
            "page_text_excerpt": page_text_excerpt,
        }

        field_texts = build_field_texts_for_matching(company_row)
        matches = score_product_matches(field_texts, product_map)
        matched_families = set()

        for match in matches:
            product_name = match["product_name"]
            family_name = match["family_name"]
            matched_keyword = match["matched_keyword"]
            source_field = match["source_field"]
            source_excerpt = match["source_excerpt"]
            match_confidence = match["match_confidence"]

            matched_families.add(family_name)

            cur.execute("SELECT id, family_id FROM products WHERE name = ?", (product_name,))
            product_row = cur.fetchone()
            if not product_row:
                continue

            product_id, family_id = product_row

            cur.execute(
                """
                INSERT OR IGNORE INTO company_products
                (company_id, product_id, match_source, source_field, matched_keyword, match_confidence)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    product_id,
                    "profile text",
                    source_field,
                    matched_keyword,
                    match_confidence,
                ),
            )

            cur.execute(
                """
                INSERT INTO product_match_audit
                (company_id, product_id, family_id, source_field, matched_keyword, source_excerpt, match_confidence)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_id,
                    product_id,
                    family_id,
                    source_field,
                    matched_keyword,
                    source_excerpt,
                    match_confidence,
                ),
            )

        for family_name in matched_families:
            cur.execute("SELECT id FROM product_families WHERE name = ?", (family_name,))
            family_row = cur.fetchone()
            if family_row:
                family_id = family_row[0]
                cur.execute(
                    """
                    INSERT OR IGNORE INTO product_family_matches (company_id, family_id)
                    VALUES (?, ?)
                    """,
                    (company_id, family_id),
                )

    sql_views = (
        "CREATE INDEX IF NOT EXISTS idx_companies_name ON companies(name);"
        "CREATE INDEX IF NOT EXISTS idx_companies_exporter_type ON companies(exporter_type);"
        "CREATE INDEX IF NOT EXISTS idx_companies_state ON companies(state);"
        "CREATE INDEX IF NOT EXISTS idx_company_products_company ON company_products(company_id);"

        "DROP VIEW IF EXISTS v_company_profile;"
        "CREATE VIEW v_company_profile AS "
        "SELECT "
        "c.id, c.name, c.exporter_type, c.website, c.address, c.state, c.postcode, "
        "c.profile_url, "
        "GROUP_CONCAT(DISTINCT cert.name) AS certifications, "
        "GROUP_CONCAT(DISTINCT acc.name) AS accreditations "
        "FROM companies c "
        "LEFT JOIN company_certifications cc ON c.id = cc.company_id "
        "LEFT JOIN certifications cert ON cc.certification_id = cert.id "
        "LEFT JOIN company_accreditations ca ON c.id = ca.company_id "
        "LEFT JOIN accreditations acc ON ca.accreditation_id = acc.id "
        "GROUP BY c.id;"

        "DROP VIEW IF EXISTS v_company_products;"
        "CREATE VIEW v_company_products AS "
        "SELECT "
        "c.id, c.name, c.exporter_type, c.website, c.state, c.profile_url, "
        "GROUP_CONCAT(DISTINCT pf.name) AS product_families, "
        "GROUP_CONCAT(DISTINCT CASE WHEN p.hierarchy_level = 2 THEN p.name END) AS product_variants "
        "FROM companies c "
        "LEFT JOIN company_products cp ON c.id = cp.company_id "
        "LEFT JOIN products p ON cp.product_id = p.id "
        "LEFT JOIN product_families pf ON p.family_id = pf.id "
        "GROUP BY c.id;"

        "DROP VIEW IF EXISTS v_product_hierarchy;"
        "CREATE VIEW v_product_hierarchy AS "
        "SELECT "
        "pf.name AS family_name, "
        "root.name AS root_product, "
        "child.name AS variant_product, "
        "child.hierarchy_level "
        "FROM product_families pf "
        "LEFT JOIN products root ON root.family_id = pf.id AND root.parent_product_id IS NULL "
        "LEFT JOIN products child ON child.parent_product_id = root.id "
        "ORDER BY pf.sort_order, child.name;"
    )

    cur.executescript(sql_views)

    conn.commit()
    conn.close()


# --------------------------------------------------
# EXPORT FROM DATABASE TO MASTER EXCEL
# --------------------------------------------------

def export_master_excel():
    conn = sqlite3.connect(DB_FILE)

    companies = pd.read_sql_query("SELECT * FROM companies ORDER BY name", conn)

    contacts = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            ct.email,
            ct.phone,
            ct.contact_source
        FROM contacts ct
        JOIN companies c ON c.id = ct.company_id
        ORDER BY c.name, ct.email, ct.phone
        """,
        conn,
    )

    certifications = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            cert.name AS certification
        FROM company_certifications cc
        JOIN companies c ON c.id = cc.company_id
        JOIN certifications cert ON cert.id = cc.certification_id
        ORDER BY c.name, cert.name
        """,
        conn,
    )

    accreditations = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            a.name AS accreditation
        FROM company_accreditations ca
        JOIN companies c ON c.id = ca.company_id
        JOIN accreditations a ON a.id = ca.accreditation_id
        ORDER BY c.name, a.name
        """,
        conn,
    )

    company_products = pd.read_sql_query("SELECT * FROM v_company_products ORDER BY name", conn)
    product_hierarchy = pd.read_sql_query("SELECT * FROM v_product_hierarchy", conn)
    company_profile = pd.read_sql_query("SELECT * FROM v_company_profile ORDER BY name", conn)

    product_match_audit = pd.read_sql_query(
        """
        SELECT
            c.name AS company_name,
            p.name AS product_name,
            pf.name AS product_family,
            a.source_field,
            a.matched_keyword,
            a.match_confidence,
            a.source_excerpt
        FROM product_match_audit a
        JOIN companies c ON c.id = a.company_id
        JOIN products p ON p.id = a.product_id
        JOIN product_families pf ON pf.id = a.family_id
        ORDER BY c.name, p.name, a.source_field
        """,
        conn,
    )

    product_summary = pd.read_sql_query(
        """
        SELECT
            pf.name AS product_family,
            COUNT(DISTINCT pfm.company_id) AS companies_with_product_family
        FROM product_family_matches pfm
        JOIN product_families pf ON pf.id = pfm.family_id
        GROUP BY pf.name
        ORDER BY companies_with_product_family DESC, pf.name
        """,
        conn,
    )

    certification_summary = pd.read_sql_query(
        """
        SELECT
            cert.name AS certification,
            COUNT(DISTINCT cc.company_id) AS companies_with_certification
        FROM company_certifications cc
        JOIN certifications cert ON cert.id = cc.certification_id
        GROUP BY cert.name
        ORDER BY companies_with_certification DESC, cert.name
        """,
        conn,
    )

    accreditation_summary = pd.read_sql_query(
        """
        SELECT
            a.name AS accreditation,
            COUNT(DISTINCT ca.company_id) AS companies_with_accreditation
        FROM company_accreditations ca
        JOIN accreditations a ON a.id = ca.accreditation_id
        GROUP BY a.name
        ORDER BY companies_with_accreditation DESC, a.name
        """,
        conn,
    )

    state_summary = pd.read_sql_query(
        """
        SELECT
            COALESCE(NULLIF(state, ''), '(blank)') AS state,
            COUNT(*) AS companies
        FROM companies
        GROUP BY COALESCE(NULLIF(state, ''), '(blank)')
        ORDER BY companies DESC, state
        """,
        conn,
    )

    exporter_type_summary = pd.read_sql_query(
        """
        SELECT
            COALESCE(NULLIF(exporter_type, ''), '(blank)') AS exporter_type,
            COUNT(*) AS companies
        FROM companies
        GROUP BY COALESCE(NULLIF(exporter_type, ''), '(blank)')
        ORDER BY companies DESC, exporter_type
        """,
        conn,
    )

    missing_data = pd.read_sql_query(
        """
        SELECT
            name,
            profile_url,
            website,
            exporter_type,
            state,
            CASE WHEN COALESCE(website, '') = '' THEN 1 ELSE 0 END AS missing_website,
            CASE WHEN NOT EXISTS (
                SELECT 1
                FROM contacts ct
                WHERE ct.company_id = companies.id
                  AND (COALESCE(ct.email, '') <> '' OR COALESCE(ct.phone, '') <> '')
            ) THEN 1 ELSE 0 END AS missing_contact,
            CASE WHEN NOT EXISTS (
                SELECT 1
                FROM company_products cp
                WHERE cp.company_id = companies.id
            ) THEN 1 ELSE 0 END AS missing_product_match
        FROM companies
        WHERE COALESCE(website, '') = ''
           OR NOT EXISTS (
                SELECT 1
                FROM contacts ct
                WHERE ct.company_id = companies.id
                  AND (COALESCE(ct.email, '') <> '' OR COALESCE(ct.phone, '') <> '')
           )
           OR NOT EXISTS (
                SELECT 1
                FROM company_products cp
                WHERE cp.company_id = companies.id
           )
        ORDER BY name
        """,
        conn,
    )

    summary = pd.DataFrame([
        {"Metric": "Generated UTC", "Value": datetime.now(UTC).isoformat(timespec="seconds")},
        {"Metric": "Companies", "Value": len(companies)},
        {"Metric": "Companies with Product Match", "Value": int(pd.read_sql_query("SELECT COUNT(DISTINCT company_id) AS n FROM company_products", conn).iloc[0, 0])},
        {"Metric": "Total Company-Product Links", "Value": int(pd.read_sql_query("SELECT COUNT(*) AS n FROM company_products", conn).iloc[0, 0])},
        {"Metric": "Companies with Certifications", "Value": int(pd.read_sql_query("SELECT COUNT(DISTINCT company_id) AS n FROM company_certifications", conn).iloc[0, 0])},
        {"Metric": "Total Company-Certification Links", "Value": int(pd.read_sql_query("SELECT COUNT(*) AS n FROM company_certifications", conn).iloc[0, 0])},
        {"Metric": "Companies with Accreditations", "Value": int(pd.read_sql_query("SELECT COUNT(DISTINCT company_id) AS n FROM company_accreditations", conn).iloc[0, 0])},
        {"Metric": "Total Company-Accreditation Links", "Value": int(pd.read_sql_query("SELECT COUNT(*) AS n FROM company_accreditations", conn).iloc[0, 0])},
        {"Metric": "Total Contact Rows", "Value": len(contacts)},
        {"Metric": "Companies Missing Website", "Value": int((companies['website'].fillna('') == '').sum())},
        {"Metric": "Companies Missing Contact Info", "Value": int(pd.read_sql_query(
            """
            SELECT COUNT(*) AS n
            FROM companies
            WHERE NOT EXISTS (
                SELECT 1
                FROM contacts ct
                WHERE ct.company_id = companies.id
                  AND (COALESCE(ct.email, '') <> '' OR COALESCE(ct.phone, '') <> '')
            )
            """,
            conn
        ).iloc[0, 0])},
        {"Metric": "Companies Missing Product Match", "Value": int(pd.read_sql_query(
            """
            SELECT COUNT(*) AS n
            FROM companies
            WHERE NOT EXISTS (
                SELECT 1
                FROM company_products cp
                WHERE cp.company_id = companies.id
            )
            """,
            conn
        ).iloc[0, 0])},
    ])

    conn.close()

    companies_for_excel = companies.copy().drop(columns=["details_json"], errors="ignore")

    if "page_text_excerpt" in companies_for_excel.columns:
        companies_for_excel["page_text_preview"] = (
            companies_for_excel["page_text_excerpt"]
            .fillna("")
            .astype(str)
            .str[:220]
        )
        companies_for_excel = companies_for_excel.drop(columns=["page_text_excerpt"], errors="ignore")

    final_excel_path, used_fallback = safe_write_excel(
        EXCEL_FILE,
        {
            "Summary": summary,
            "Product Summary": product_summary,
            "Certification Summary": certification_summary,
            "Accreditation Summary": accreditation_summary,
            "State Summary": state_summary,
            "Exporter Type Summary": exporter_type_summary,
            "Missing Data": missing_data,
            "Company Profile": company_profile,
            "Company Products": company_products,
            "Companies": companies_for_excel,
            "Contacts": contacts,
            "Product Match Audit": product_match_audit,
            "Certifications": certifications,
            "Accreditations": accreditations,
            "Product Hierarchy": product_hierarchy,
        },
    )

    dashboard_polish_workbook(final_excel_path)
    return final_excel_path, used_fallback


# --------------------------------------------------
# SCHEMA DOCS / RUNNER
# --------------------------------------------------

def export_schema_files():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(
        "SELECT type, name, sql FROM sqlite_master "
        "WHERE sql IS NOT NULL AND name != 'sqlite_sequence' "
        "ORDER BY type, name"
    )
    rows = cur.fetchall()

    sql_lines = [row[2] + ";" for row in rows if row[2]]
    Path(SCHEMA_SQL_FILE).write_text("\n\n".join(sql_lines), encoding="utf-8")

    md_parts = [
        "# Exporters Database Schema",
        "",
        f"Generated: {datetime.now(UTC).isoformat(timespec='seconds')}",
        "",
        "## Overview",
        "This project stores exporter listings, contacts, certifications, accreditations, products, and reporting views in exporters_final.db.",
        "",
        "## Main Reporting Views",
        "- **v_company_profile**: one row per company with certifications and accreditations rolled up.",
        "- **v_company_products**: one row per company with product family and product variant rollups.",
        "- **v_product_hierarchy**: product family hierarchy reference.",
        "",
        "## Known Limitations",
        "- Product matching is keyword based, not ML-based.",
        "- Contacts are extracted from visible text and links, so some false positives/negatives can still occur.",
        "- Some exporter pages contain repeated site boilerplate which is filtered on a best-effort basis.",
        "",
    ]

    purpose_map = {
        "companies": "Master exporter/company records.",
        "contacts": "Extracted contact details by company.",
        "attributes": "Raw extracted text attributes kept for traceability.",
        "certifications": "Certification lookup table.",
        "company_certifications": "Many-to-many link between companies and certifications.",
        "accreditations": "Accreditation lookup table.",
        "company_accreditations": "Many-to-many link between companies and accreditations.",
        "product_families": "Top-level product family dimension.",
        "products": "Product dimension with hierarchy metadata.",
        "company_products": "Matched company-to-product relationships.",
        "product_family_matches": "Matched company-to-product-family relationships.",
        "product_match_audit": "Audit table showing why product matches were made.",
        "v_company_profile": "Profile reporting view.",
        "v_company_products": "Company-product reporting view.",
        "v_product_hierarchy": "Hierarchy reference view.",
    }

    for obj_type, name, sql in rows:
        md_parts.append(f"## {obj_type.title()}: {name}")
        md_parts.append("")
        if name in purpose_map:
            md_parts.append(purpose_map[name])
            md_parts.append("")
        md_parts.append("```sql")
        md_parts.append(sql + ";")
        md_parts.append("```")
        md_parts.append("")

    Path(SCHEMA_MD_FILE).write_text("\n".join(md_parts), encoding="utf-8")
    conn.close()


def write_run_bat():
    content = (
        "@echo off\n"
        "cd /d \"%~dp0\"\n"
        "python -m py_compile pipeline_final.py\n"
        "if errorlevel 1 goto :end\n"
        "python -u pipeline_final.py\n"
        ":end\n"
        "pause\n"
    )
    Path(RUN_BAT_FILE).write_text(content, encoding="utf-8")


# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main():
    print("Collecting listings...")
    listings = collect_listings()
    print(f"Collected {len(listings)} rows")

    print("Collecting profile details...")
    enriched = enrich_rows(listings)
    df = clean_dataframe(pd.DataFrame(enriched))

    print("Building database...")
    build_database(df)

    print("Exporting Excel from database...")
    final_excel_path, used_fallback = export_master_excel()

    export_schema_files()
    write_run_bat()
    cleanup_fallback_excels_if_master_written(used_fallback)

    print("FINAL PIPELINE COMPLETE")
    print("Created:")
    print(f"- {DB_FILE}")
    print(f"- {final_excel_path}")
    print(f"- {SCHEMA_SQL_FILE}")
    print(f"- {SCHEMA_MD_FILE}")
    print(f"- {RUN_BAT_FILE}")

    if used_fallback:
        print("exporters_master.xlsx was locked (likely open in Excel).")
        print("A fallback Excel file was created.")
        print("It will automatically replace exporters_master.xlsx after Excel is closed.")
    else:
        print("Database, Excel, schema files, and batch runner are fully in sync.")


if __name__ == "__main__":
    if len(sys.argv) >= 4 and sys.argv[1] == "--finalize-fallback":
        finalize_locked_excel(sys.argv[2], sys.argv[3])
        sys.exit(0)

    main()
